"""
The Flourish Collective — Financial Dashboard
Reads from Flourish_Financial_Tracker.xlsx
Pages: Executive Summary | FY26 Budget vs Actual | Revenue Detail & Trends | Functional Expense Report
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
import os

st.set_page_config(
    page_title="The Flourish Collective",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Brand ───────────────────────────────────────────────────
GREEN      = "#3b5236"
GOLD       = "#ecd274"
CREAM      = "#f5f2ea"
DARK_GREEN = "#2a3d27"
SLATE      = "#455A64"
TERRACOTTA = "#c84c33"
NAVY       = "#184c7c"
PURPLE     = "#6f1468"
BROWN      = "#724b3b"
WHITE      = "#ffffff"

CAT_COLORS = {
    "Development: Donor & Partner Care":        NAVY,
    "Program: Education, Allyship, Engagement": GREEN,
    "Overhead":                                  SLATE,
    "Grant Making":                              GOLD,
    "Split per JD allocations":                 BROWN,
}

REV_COLORS = {
    "Community Contributions":       GREEN,
    "Fundraising Events":            TERRACOTTA,
    "Learning Events":               NAVY,
    "Annual Gathering":              PURPLE,
    "Merchandise":                   BROWN,
    "Reimbursables/Misc":           SLATE,
}

st.markdown(f"""
<link href="https://fonts.googleapis.com/css2?family=Lora:ital,wght@0,400;0,600;0,700;1,400&family=Darker+Grotesque:wght@400;500;600;700;900&display=swap" rel="stylesheet">
<style>
  .stApp {{ background-color:{CREAM}; }}
  section[data-testid="stSidebar"] {{
    background-color:{GREEN} !important;
    border-right: 3px solid {GOLD};
  }}
  section[data-testid="stSidebar"] * {{ color:{CREAM} !important; }}
  section[data-testid="stSidebar"] .stRadio label {{
    font-family:'Darker Grotesque',sans-serif;
    font-size:15px; font-weight:600;
  }}
  section[data-testid="stSidebar"] hr {{ border-color:rgba(236,210,116,.35) !important; }}
  h1,h2,h3,h4 {{ font-family:'Lora',serif !important; color:{GREEN} !important; }}
  p,li,span,div,label,td,th {{ font-family:'Darker Grotesque',sans-serif !important; }}
  .stMetric label {{
    font-family:'Darker Grotesque',sans-serif !important;
    font-weight:700 !important; color:{GREEN} !important;
    font-size:11px !important; text-transform:uppercase; letter-spacing:.07em;
  }}
  [data-testid="stMetricValue"] {{
    font-family:'Lora',serif !important; color:{GREEN} !important;
    font-size:1.65rem !important; font-weight:700 !important;
  }}
  [data-testid="stMetricDelta"] {{ font-family:'Darker Grotesque',sans-serif !important; }}
  [data-testid="metric-container"] {{
    background:{WHITE}; border:1px solid rgba(59,82,54,.1);
    border-top:3px solid {GREEN}; border-radius:10px;
    padding:1rem 1.2rem !important;
    box-shadow:0 2px 8px rgba(59,82,54,.06);
  }}
  .hero {{
    background:linear-gradient(135deg,{GREEN} 0%,{DARK_GREEN} 60%,#1a2618 100%);
    border-radius:12px; padding:2.2rem 2.8rem; margin-bottom:1.8rem;
    border-bottom:4px solid {GOLD};
  }}
  .hero h1 {{ color:{CREAM} !important; font-size:1.9rem; margin:0 0 .3rem; }}
  .hero p {{ color:rgba(245,242,234,.7) !important; font-size:.95rem; margin:0; }}
  .hero-tag {{
    color:{GOLD} !important; font-size:.72rem !important;
    text-transform:uppercase; letter-spacing:.14em;
    font-weight:700; margin-bottom:.4rem !important;
  }}
  .section-head {{
    background:{GREEN}; border-radius:8px; padding:.9rem 1.4rem;
    margin:1.4rem 0 .8rem; border-left:5px solid {GOLD};
  }}
  .section-head h3 {{ color:{CREAM} !important; margin:0; font-size:1.1rem; }}
  .section-head p {{ color:rgba(245,242,234,.65) !important; margin:.2rem 0 0; font-size:.84rem; }}
  .note {{
    background:rgba(59,82,54,.06); border-left:3px solid {GOLD};
    border-radius:0 8px 8px 0; padding:.7rem 1rem;
    font-size:.87rem; color:{GREEN}; margin:.8rem 0;
  }}
</style>
""", unsafe_allow_html=True)


# ── Data Loading ────────────────────────────────────────────
@st.cache_data(ttl=300)
def load_data():
    path = "Flourish_Financial_Tracker.xlsx"
    if not os.path.exists(path):
        return None

    wb = openpyxl.load_workbook(path, data_only=True)

    def fv(ws, r, c):
        val = ws.cell(r, c).value
        if val is None: return 0.0
        try: return float(val)
        except: return 0.0

    def sv(ws, r, c):
        val = ws.cell(r, c).value
        return str(val).strip() if val else ""

    # ── Progress Tracker ─────────────────────────────────────
    pt = wb["Progress Tracker"]
    MON = list(range(8, 20))  # H–S Jan–Dec

    def pt_ytd(row):  return sum(fv(pt, row, c) for c in MON)
    def pt_bgt(row):
        b = fv(pt, row, 4)
        if b == 0: b = fv(pt, row, 5)
        return b

    rev_rows = {
        "Executive Board Contributions":       7,
        "Leadership Board Contributions":      8,
        "Founding Donor (NFF) Contributions":  9,
        "Major Donor Contributions":           10,
        "All Other Community Contributions":   11,
        "Merchandise":                         12,
        "Annual Gathering":                    13,
        "Gathering Scholarships":              14,
        "Fundraising Events":                  15,
        "Online Learning Events":              17,
        "In Person Learning Events":           18,
        "Reimbursables/Misc":                 19,
    }
    rev_budget = {k: pt_bgt(r) for k, r in rev_rows.items()}
    rev_ytd    = {k: pt_ytd(r) for k, r in rev_rows.items()}
    total_rev_budget = sum(rev_budget.values())
    total_rev_ytd    = sum(rev_ytd.values())

    exp_rows = {
        "Salaries & Taxes":                     (23, None),
        "Gusto - Payroll Fees":                 (24, None),
        "Medicare & SS Deductions":             (25, None),
        "SSB Trust (401K)":                     (26, None),
        "Guideline (401K Fees)":                (27, None),
        "Next Insurance (Worker's Comp)":       (28, None),
        "Shelterpoint (NY FL & SDI)":           (29, None),
        "Health Insurance":                     (30, None),
        "Beam - HSA Fees":                      (31, None),
        "Member Care":                          (32, None),
        "Impact Partner Care":                  (33, None),
        "Fundraising & Marketing Expenses":     (34, None),
        "Fundraising Event":                    (35, None),
        "Sales Tax (Fundraising Event)":        (36, None),
        "Bank & Donation Fees":                 (37, None),
        "Subscription Services":                (38, None),
        "Leadership Travel":                    (39, None),
        "Leadership Meetings":                  (40, None),
        "Postage and Mailing Supplies":         (41, None),
        "Non-Profit Liability Ins":             (42, None),
        "Supplies":                             (43, None),
        "In Person Annual Event":               (44, None),
        "Training":                             (45, None),
        "Honorariums":                          (46, None),
        "Learning Events Expenses":             (47, None),
        "Merchandise (Expense)":                (48, None),
        "Misc Expenses":                        (49, None),
        "Printing":                             (50, None),
        "Grants":                               (51, None),
        "Kristin Malvadkar":                    (52, None),
        "Lauren (Communications)":              (53, None),
        "Financial Services":                   (54, None),
        "Sharon Bennett (Admin)":               (55, None),
        "Contractor Support (PM/DA)":           (56, None),
    }
    exp_category = {}
    exp_budget   = {}
    exp_ytd_d    = {}
    for name, (row, _) in exp_rows.items():
        exp_category[name] = sv(pt, row, 3)
        exp_budget[name]   = fv(pt, row, 4)
        exp_ytd_d[name]    = pt_ytd(row)

    total_exp_budget = sum(exp_budget.values())
    total_exp_ytd    = sum(exp_ytd_d.values())
    net_income_budget = total_rev_budget - total_exp_budget
    net_income_ytd    = total_rev_ytd    - total_exp_ytd

    # ── Functional Expense Report ─────────────────────────────
    fer = wb["Functional Expense Report"]
    func_cats = [
        "Development: Donor & Partner Care",
        "Program: Education, Allyship, Engagement",
        "Overhead",
        "Grant Making",
    ]
    func_budget = {c: fv(fer, r, 3) for c, r in zip(func_cats, [43,44,45,46])}
    func_ytd    = {c: fv(fer, r, 4) for c, r in zip(func_cats, [43,44,45,46])}

    # ── Revenue Trends ────────────────────────────────────────
    rt = wb["Revenue Trends"]
    trend_sources = [sv(rt, r, 1) for r in range(5, 14) if sv(rt, r, 1)]
    trend_data = []
    for i, src in enumerate(trend_sources):
        row = 5 + i
        for pi, period in enumerate(["Q1","Q2","Q3","Q4","Annual"]):
            trend_data.append({
                "Source": src, "Period": period,
                "FY24": fv(rt, row, 2+pi*3),
                "FY25": fv(rt, row, 3+pi*3),
                "FY26": fv(rt, row, 4+pi*3),
            })
    for pi, period in enumerate(["Q1","Q2","Q3","Q4","Annual"]):
        trend_data.append({
            "Source": "TOTAL REVENUE", "Period": period,
            "FY24": fv(rt, 14, 2+pi*3),
            "FY25": fv(rt, 14, 3+pi*3),
            "FY26": fv(rt, 14, 4+pi*3),
        })

    # ── Revenue Detail historical ─────────────────────────────
    rd = wb["Revenue Detail"]
    hist_rev_rows = [
        (6,  "Community Contributions", "parent"),
        (12, "Merchandise",             "parent"),
        (13, "Annual Gathering",        "parent"),
        (15, "Fundraising Events",      "parent"),
        (16, "Learning Events",         "parent"),
        (19, "Reimbursables/Misc",      "parent"),
        (20, "TOTAL REVENUE",           "total"),
    ]
    hist_rev = []
    for row, name, level in hist_rev_rows:
        for yi, yr in enumerate(["FY21","FY22","FY23","FY24","FY25"]):
            hist_rev.append({
                "Account": name, "Year": yr, "Level": level,
                "Amount": fv(rd, row, 3+yi),
            })

    # ── Executive Dashboard historical ───────────────────────
    ex = wb["Executive Dashboard"]
    YEARS = ["FY21","FY22","FY23","FY24","FY25"]
    hist_summary = {
        "Total Revenue":  [fv(ex, 12, c) for c in range(2,7)],
        "Total Expenses": [fv(ex, 18, c) for c in range(2,7)],
        "Net Income":     [fv(ex, 20, c) for c in range(2,7)],
    }
    hist_community = {
        "Active Allies": [fv(ex, 25, c) for c in range(2,7)],
        "Donors":        [fv(ex, 26, c) for c in range(2,7)],
        "Avg Donation":  [fv(ex, 27, c) for c in range(2,7)],
    }

    return {
        "rev_budget": rev_budget, "rev_ytd": rev_ytd,
        "total_rev_budget": total_rev_budget, "total_rev_ytd": total_rev_ytd,
        "exp_budget": exp_budget, "exp_ytd": exp_ytd_d,
        "exp_category": exp_category,
        "total_exp_budget": total_exp_budget, "total_exp_ytd": total_exp_ytd,
        "net_income_budget": net_income_budget, "net_income_ytd": net_income_ytd,
        "trend_data": trend_data,
        "hist_rev": hist_rev,
        "hist_summary": hist_summary, "hist_community": hist_community,
        "hist_years": YEARS,
    }


def fmt(val):
    try:
        val = float(val)
        if val < 0: return f"-${abs(val):,.0f}"
        return f"${val:,.0f}"
    except: return "—"

def progress_bar(ytd, budget, color=GREEN):
    p = min(ytd / budget, 1.2) if budget > 0 else 0
    disp = ytd / budget if budget > 0 else 0
    bar_c = TERRACOTTA if disp > 1.05 else color
    return f"""<div style="margin:2px 0 8px">
      <div style="background:rgba(59,82,54,.1);border-radius:4px;height:7px;overflow:hidden">
        <div style="width:{min(p*100,100):.1f}%;height:100%;background:{bar_c};border-radius:4px"></div>
      </div>
      <span style="font-size:.74rem;color:{SLATE}">{disp:.1%} of budget</span>
    </div>"""

def section(title, subtitle=""):
    sub = f"<p>{subtitle}</p>" if subtitle else ""
    st.markdown(f'<div class="section-head"><h3>{title}</h3>{sub}</div>',
                unsafe_allow_html=True)


# ── Sidebar ──────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"""
    <div style="padding:.5rem 0 1rem">
      <div style="font-size:.68rem;text-transform:uppercase;letter-spacing:.14em;
                  color:rgba(236,210,116,.8);font-weight:700;margin-bottom:.2rem">
        The Flourish Collective
      </div>
      <div style="font-family:'Lora',serif;font-size:1.25rem;font-weight:700;
                  color:#f5f2ea;line-height:1.3">
        Financial Dashboard
      </div>
    </div>""", unsafe_allow_html=True)
    st.markdown("---")

    page = st.radio("Navigate", [
        "📊 Executive Summary",
        "📋 FY26 Budget vs. Actual",
        "📈 Revenue Detail & Trends",
    ], label_visibility="collapsed")

    st.markdown("---")
    if st.button("🔄 Refresh Data"):
        st.cache_data.clear()
        st.rerun()

    st.markdown(f"""
    <div style="font-size:.75rem;color:rgba(245,242,234,.45);
                margin-top:1rem;line-height:1.6">
      FY26 data live from tracker.<br>
      FY21–FY25 historical.<br><br>
      <em>Building allies &amp;<br>investing in leaders<br>for racial justice.</em>
    </div>""", unsafe_allow_html=True)


# ── Load ─────────────────────────────────────────────────────
data = load_data()
if data is None:
    st.error("⚠️ Flourish_Financial_Tracker.xlsx not found in repository root.")
    st.stop()
d = data


# ════════════════════════════════════════════════════════════
# PAGE 1 — EXECUTIVE SUMMARY
# ════════════════════════════════════════════════════════════
if page == "📊 Executive Summary":
    st.markdown(f"""<div class="hero">
      <div class="hero-tag">FY26 Year-to-Date</div>
      <h1>Executive Summary</h1>
      <p>High-level financial performance and community growth — Q1 complete</p>
    </div>""", unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    rev_pct = d["total_rev_ytd"] / d["total_rev_budget"] if d["total_rev_budget"] else 0
    exp_pct = d["total_exp_ytd"] / d["total_exp_budget"] if d["total_exp_budget"] else 0
    with c1:
        st.metric("Total Revenue YTD", fmt(d["total_rev_ytd"]),
                  f"{rev_pct:.1%} of {fmt(d['total_rev_budget'])} budget")
    with c2:
        st.metric("Total Expenses YTD", fmt(d["total_exp_ytd"]),
                  f"{exp_pct:.1%} of {fmt(d['total_exp_budget'])} budget")
    with c3:
        ni = d["net_income_ytd"]
        st.metric("Net Income YTD", fmt(ni), "surplus" if ni >= 0 else "deficit")
    with c4:
        st.metric("FY26 Budgeted Net Income", fmt(d["net_income_budget"]))

    st.markdown(f"<div class='note'>📅 <strong>Q1 complete (Jan–Mar).</strong> Q2–Q4 will populate as monthly actuals are entered into the Progress Tracker.</div>", unsafe_allow_html=True)

    section("Historical Financial Performance", "FY21–FY25 actuals")

    years = d["hist_years"]
    hs    = d["hist_summary"]

    fig = go.Figure()
    fig.add_bar(name="Total Revenue", x=years, y=hs["Total Revenue"],
                marker_color=GREEN, opacity=.85)
    fig.add_bar(name="Total Expenses", x=years, y=hs["Total Expenses"],
                marker_color=TERRACOTTA, opacity=.85)
    fig.add_scatter(name="Net Income", x=years, y=hs["Net Income"],
                    mode="lines+markers",
                    line=dict(color=GOLD, width=3),
                    marker=dict(size=8, color=GOLD, line=dict(color=DARK_GREEN, width=1.5)))
    fig.update_layout(
        barmode="group", plot_bgcolor="white", paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Darker Grotesque", size=12),
        legend=dict(orientation="h", y=1.1, x=0),
        margin=dict(l=0,r=0,t=30,b=0), height=320,
        yaxis=dict(tickprefix="$", tickformat=",.0f", gridcolor="#ece9e1"),
        xaxis=dict(gridcolor="rgba(0,0,0,0)"),
    )
    st.plotly_chart(fig, use_container_width=True)

    col1, col2 = st.columns([3, 2])

    with col1:
        section("FY26 Revenue Snapshot", "YTD vs. full-year budget")
        key_rev = [
            ("Founding Donor (NFF)",    "Founding Donor (NFF) Contributions"),
            ("All Other Community",     "All Other Community Contributions"),
            ("Fundraising Events",      "Fundraising Events"),
            ("Online Learning Events",  "Online Learning Events"),
            ("Executive Board",         "Executive Board Contributions"),
            ("Leadership Board",        "Leadership Board Contributions"),
        ]
        for label, key in key_rev:
            bgt = d["rev_budget"].get(key, 0)
            ytd = d["rev_ytd"].get(key, 0)
            if bgt > 0:
                st.markdown(f"**{label}** — {fmt(ytd)} / {fmt(bgt)}")
                st.markdown(progress_bar(ytd, bgt), unsafe_allow_html=True)

    with col2:
        section("Community Growth", "Historical trends")
        hc = d["hist_community"]
        fig2 = go.Figure()
        fig2.add_scatter(x=years, y=hc["Active Allies"], name="Active Allies",
                         mode="lines+markers",
                         line=dict(color=GREEN, width=2.5),
                         marker=dict(size=7))
        fig2.add_scatter(x=years, y=hc["Donors"], name="# Donors",
                         mode="lines+markers",
                         line=dict(color=NAVY, width=2.5),
                         marker=dict(size=7))
        fig2.add_hline(y=330, line_dash="dash", line_color=GOLD, line_width=1.5,
                       annotation_text="330 target", annotation_font_size=10)
        fig2.update_layout(
            plot_bgcolor="white", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Darker Grotesque", size=11),
            legend=dict(orientation="h", y=1.12, x=0),
            margin=dict(l=0,r=0,t=30,b=0), height=240,
            yaxis=dict(gridcolor="#ece9e1"),
            xaxis=dict(gridcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig2, use_container_width=True)
        m1, m2, m3 = st.columns(3)
        with m1: st.metric("FY25 Allies", f"{int(hc['Active Allies'][-1])}", "→ 330")
        with m2: st.metric("FY25 Donors", f"{int(hc['Donors'][-1])}", "→ 198")
        with m3: st.metric("Avg Gift", fmt(hc["Avg Donation"][-1]))


# ════════════════════════════════════════════════════════════
# PAGE 2 — FY26 BUDGET VS ACTUAL
# ════════════════════════════════════════════════════════════
elif page == "📋 FY26 Budget vs. Actual":
    st.markdown(f"""<div class="hero">
      <div class="hero-tag">FY26 Current Year</div>
      <h1>Budget vs. Actual</h1>
      <p>All revenue and expense accounts — Q1 actuals vs. full-year budget</p>
    </div>""", unsafe_allow_html=True)

    tab_rev, tab_exp = st.tabs(["💰 Revenue", "💸 Expenses"])

    with tab_rev:
        section("Revenue — FY26 Budget vs. YTD Actual")

        rev_items = [
            ("Executive Board Contributions",       True),
            ("Leadership Board Contributions",      True),
            ("Founding Donor (NFF) Contributions",  True),
            ("Major Donor Contributions",           True),
            ("All Other Community Contributions",   True),
            ("Merchandise",                         False),
            ("Annual Gathering",                    False),
            ("Gathering Scholarships",              False),
            ("Fundraising Events",                  False),
            ("Online Learning Events",              True),
            ("In Person Learning Events",           True),
            ("Reimbursables/Misc",                  False),
        ]

        rows = []
        for key, indent in rev_items:
            bgt = d["rev_budget"].get(key, 0)
            ytd = d["rev_ytd"].get(key, 0)
            rows.append({
                "Account": ("  " if indent else "") + key,
                "FY26 Budget": bgt, "YTD Actual": ytd,
                "Remaining": bgt - ytd,
                "% of Budget": ytd/bgt if bgt > 0 else 0,
            })
        rows.append({
            "Account": "TOTAL REVENUE",
            "FY26 Budget": d["total_rev_budget"],
            "YTD Actual": d["total_rev_ytd"],
            "Remaining": d["total_rev_budget"] - d["total_rev_ytd"],
            "% of Budget": d["total_rev_ytd"] / d["total_rev_budget"] if d["total_rev_budget"] else 0,
        })
        df_rev = pd.DataFrame(rows)

        chart_df = df_rev[~df_rev["Account"].str.contains("TOTAL")].copy()
        chart_df["Account"] = chart_df["Account"].str.strip()
        chart_df = chart_df[chart_df["FY26 Budget"] > 0]

        fig = go.Figure()
        fig.add_bar(name="FY26 Budget", x=chart_df["Account"], y=chart_df["FY26 Budget"],
                    marker_color=f"rgba(59,82,54,.2)",
                    marker_line_color=GREEN, marker_line_width=1.5)
        fig.add_bar(name="YTD Actual", x=chart_df["Account"], y=chart_df["YTD Actual"],
                    marker_color=GREEN)
        fig.update_layout(
            barmode="overlay", plot_bgcolor="white", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Darker Grotesque", size=11),
            legend=dict(orientation="h", y=1.08, x=0),
            margin=dict(l=0,r=0,t=30,b=90), height=320,
            yaxis=dict(tickprefix="$", tickformat=",.0f", gridcolor="#ece9e1"),
            xaxis=dict(tickangle=-35, gridcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig, use_container_width=True)

        st.dataframe(
            df_rev.style
                .format({"FY26 Budget":"${:,.0f}","YTD Actual":"${:,.0f}",
                         "Remaining":"${:,.0f}","% of Budget":"{:.1%}"})
                .apply(lambda x: [
                    f"background:{CREAM};font-weight:700"
                    if x["Account"] == "TOTAL REVENUE" else "" for _ in x], axis=1),
            use_container_width=True, hide_index=True
        )

    with tab_exp:
        section("Expenses — FY26 Budget vs. YTD Actual")

        cat_order = [
            "Split per JD allocations",
            "Development: Donor & Partner Care",
            "Program: Education, Allyship, Engagement",
            "Overhead",
            "Grant Making",
        ]

        rows_exp = []
        for cat in cat_order:
            cat_bgt = sum(b for n,b in d["exp_budget"].items() if d["exp_category"].get(n)==cat)
            cat_ytd = sum(y for n,y in d["exp_ytd"].items()    if d["exp_category"].get(n)==cat)
            rows_exp.append({
                "Account": f"▸ {cat}", "FY26 Budget": cat_bgt, "YTD Actual": cat_ytd,
                "Remaining": cat_bgt-cat_ytd,
                "% of Budget": cat_ytd/cat_bgt if cat_bgt > 0 else 0,
                "_grp": True, "_cat": cat,
            })
            for name in sorted(d["exp_budget"]):
                if d["exp_category"].get(name) == cat and d["exp_budget"][name] > 0:
                    bgt = d["exp_budget"][name]; ytd = d["exp_ytd"].get(name, 0)
                    rows_exp.append({
                        "Account": f"  {name}", "FY26 Budget": bgt, "YTD Actual": ytd,
                        "Remaining": bgt-ytd,
                        "% of Budget": ytd/bgt if bgt > 0 else 0,
                        "_grp": False, "_cat": cat,
                    })
        rows_exp.append({
            "Account": "TOTAL EXPENSES",
            "FY26 Budget": d["total_exp_budget"], "YTD Actual": d["total_exp_ytd"],
            "Remaining": d["total_exp_budget"]-d["total_exp_ytd"],
            "% of Budget": d["total_exp_ytd"]/d["total_exp_budget"] if d["total_exp_budget"] else 0,
            "_grp": True, "_cat": "",
        })
        df_exp = pd.DataFrame(rows_exp)

        # Category chart
        cat_df = df_exp[df_exp["_grp"] & (df_exp["Account"] != "TOTAL EXPENSES")].copy()
        cat_df["Label"] = cat_df["Account"].str.replace("▸ ", "")
        fig2 = go.Figure()
        for _, row_c in cat_df.iterrows():
            color = CAT_COLORS.get(row_c["_cat"], SLATE)
            fig2.add_bar(name=row_c["Label"], x=[row_c["Label"]],
                         y=[row_c["FY26 Budget"]],
                         marker_color=color, opacity=.25,
                         marker_line_color=color, marker_line_width=1.5,
                         showlegend=False)
            fig2.add_bar(name=row_c["Label"]+" YTD", x=[row_c["Label"]],
                         y=[row_c["YTD Actual"]],
                         marker_color=color, showlegend=False)
        fig2.update_layout(
            barmode="overlay", plot_bgcolor="white", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Darker Grotesque", size=11),
            margin=dict(l=0,r=0,t=10,b=70), height=280,
            yaxis=dict(tickprefix="$", tickformat=",.0f", gridcolor="#ece9e1"),
            xaxis=dict(tickangle=-20, gridcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig2, use_container_width=True)

        show_cols = ["Account","FY26 Budget","YTD Actual","Remaining","% of Budget"]
        st.dataframe(
            df_exp[show_cols].style
                .format({"FY26 Budget":"${:,.0f}","YTD Actual":"${:,.0f}",
                         "Remaining":"${:,.0f}","% of Budget":"{:.1%}"})
                .apply(lambda x: [
                    f"background:{CREAM};font-weight:700"
                    if "TOTAL" in x["Account"] or x["Account"].startswith("▸")
                    else "" for _ in x], axis=1),
            use_container_width=True, hide_index=True
        )


# ════════════════════════════════════════════════════════════
# PAGE 3 — REVENUE DETAIL & TRENDS
# ════════════════════════════════════════════════════════════
elif page == "📈 Revenue Detail & Trends":
    st.markdown(f"""<div class="hero">
      <div class="hero-tag">Revenue Analysis</div>
      <h1>Revenue Detail & Trends</h1>
      <p>Historical revenue FY21–FY25 · Quarterly comparison FY24/FY25/FY26</p>
    </div>""", unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["📊 Historical (FY21–FY25)", "📅 Quarterly Comparison (FY24–FY26)"])

    with tab1:
        section("Revenue by Source — FY21 to FY25",
                "Income statement source — pattern analysis only")

        df_hist = pd.DataFrame(d["hist_rev"])
        df_plot = df_hist[(df_hist["Level"]=="parent") & (df_hist["Account"]!="TOTAL REVENUE") & (df_hist["Amount"]>0)]

        fig = px.bar(
            df_plot, x="Year", y="Amount", color="Account",
            color_discrete_map={a: REV_COLORS.get(a, SLATE) for a in df_plot["Account"].unique()},
            barmode="stack",
        )
        fig.update_layout(
            plot_bgcolor="white", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Darker Grotesque", size=12),
            legend=dict(orientation="h", y=-.2, x=0),
            margin=dict(l=0,r=0,t=20,b=100), height=360,
            yaxis=dict(tickprefix="$", tickformat=",.0f", gridcolor="#ece9e1"),
            xaxis=dict(gridcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig, use_container_width=True)

        df_total = df_hist[df_hist["Account"]=="TOTAL REVENUE"]
        fig2 = go.Figure()
        fig2.add_scatter(
            x=df_total["Year"], y=df_total["Amount"],
            mode="lines+markers+text",
            line=dict(color=GREEN, width=3),
            marker=dict(size=9, color=GREEN, line=dict(color=DARK_GREEN, width=1.5)),
            text=[fmt(v) for v in df_total["Amount"]],
            textposition="top center",
            textfont=dict(size=11, family="Darker Grotesque"),
        )
        fig2.update_layout(
            title="Total Revenue Trend FY21–FY25",
            plot_bgcolor="white", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Darker Grotesque", size=12),
            margin=dict(l=0,r=0,t=40,b=0), height=200,
            yaxis=dict(tickprefix="$", tickformat=",.0f", gridcolor="#ece9e1"),
            xaxis=dict(gridcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig2, use_container_width=True)

        # Historical pivot table
        df_pivot = df_plot.pivot_table(index="Account", columns="Year", values="Amount", fill_value=0)
        st.markdown("**Revenue by Account — FY21 to FY25**")
        st.dataframe(
            df_pivot.style.format("${:,.0f}"),
            use_container_width=True
        )

    with tab2:
        section("Quarterly Revenue Comparison — FY24 / FY25 / FY26",
                "FY24 & FY25 income statement · FY26 live from Progress Tracker")

        df_trend = pd.DataFrame(d["trend_data"])
        df_tot = df_trend[df_trend["Source"]=="TOTAL REVENUE"]

        # Total quarterly grouped bar
        fig3 = go.Figure()
        for yr, color in [("FY24", NAVY), ("FY25", TERRACOTTA), ("FY26", GREEN)]:
            vals = []
            for p in ["Q1","Q2","Q3","Q4"]:
                row = df_tot[df_tot["Period"]==p]
                vals.append(float(row[yr].values[0]) if len(row) > 0 else 0)
            fig3.add_bar(name=yr, x=["Q1","Q2","Q3","Q4"], y=vals,
                         marker_color=color, opacity=.85)
        fig3.update_layout(
            barmode="group", title="Total Revenue by Quarter",
            plot_bgcolor="white", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Darker Grotesque", size=12),
            legend=dict(orientation="h", y=1.1, x=0),
            margin=dict(l=0,r=0,t=50,b=0), height=300,
            yaxis=dict(tickprefix="$", tickformat=",.0f", gridcolor="#ece9e1"),
            xaxis=dict(gridcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig3, use_container_width=True)

        # Quarterly table
        periods  = ["Q1","Q2","Q3","Q4","Annual"]
        sources  = [s for s in df_trend["Source"].unique() if s != "TOTAL REVENUE"]
        rows_t   = []
        for src in sources:
            row_d = {"Revenue Source": src}
            src_df = df_trend[df_trend["Source"]==src]
            for p in periods:
                p_df = src_df[src_df["Period"]==p]
                for yr in ["FY24","FY25","FY26"]:
                    val = float(p_df[yr].values[0]) if len(p_df)>0 else 0
                    row_d[f"{p} {yr}"] = val if val else None
            rows_t.append(row_d)

        tot_r = {"Revenue Source": "TOTAL REVENUE"}
        for p in periods:
            p_df = df_tot[df_tot["Period"]==p]
            for yr in ["FY24","FY25","FY26"]:
                val = float(p_df[yr].values[0]) if len(p_df)>0 else 0
                tot_r[f"{p} {yr}"] = val if val else None
        rows_t.append(tot_r)

        df_table = pd.DataFrame(rows_t)
        num_cols = [c for c in df_table.columns if c != "Revenue Source"]
        st.dataframe(
            df_table.style
                .format({c:"${:,.0f}" for c in num_cols}, na_rep="—")
                .apply(lambda x: [
                    f"background:{CREAM};font-weight:700"
                    if x["Revenue Source"]=="TOTAL REVENUE" else "" for _ in x], axis=1),
            use_container_width=True, hide_index=True
        )
        st.markdown(f"<div class='note'>⚠️ FY24 & FY25 from income statement for pattern analysis only — not official reporting. FY26 Q2–Q4 populates as monthly actuals are entered.</div>", unsafe_allow_html=True)
